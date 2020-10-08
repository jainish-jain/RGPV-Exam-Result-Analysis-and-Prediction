from selenium import webdriver
import time
import csv, xlsxwriter 
#import cv2
import os ,stat, sys
import requests
import PIL.Image
import pandas as pd
import pymysql
from PIL import Image,ImageFile
#ImageFile.LOAD_TRUNCATED_IMAGES = True
import pytesseract
from selenium.webdriver.chrome.options import Options
import tkinter as tk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinter import *
from tkinter import messagebox
from webdriver_manager.chrome import ChromeDriverManager

from tkinter.font import Font
from shutil import copyfile
import openpyxl
from tkinter.simpledialog import askstring
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from matplotlib import pyplot as plt
from tkinter import filedialog,ttk,Canvas
from tkinter.filedialog import askopenfile
from sklearn.ensemble import RandomForestRegressor 
from sklearn.model_selection import train_test_split 
from matplotlib.animation import FuncAnimation
from PIL import ImageTk, Image
#pytesseract.pytesseract.tesseract_cmd = r'C:\Users\jaini\AppData\Local\Tesseract-OCR\tesseract.exe'
from cv2 import cv2
# Create instance of FieldStorage 
#form = cgi.FieldStorage() 





# Get data from fields
# frno = form.getvalue('frno')
# lrno  = form.getvalue('lrno')
# count_sub= form.getvalue('num_sub')
# course=form.getvalue('course')
# sem =form.getvalue('sem')
'''
count_sub="11"
course="btech"
sem="5"
rno=['0808cs171071','0808cs171072','0808cs171073','0808cs171074','0808cs171075','0808cs171076','0808cs171077','0808cs171078','0808cs171079']
'''
'''    
db = pymysql.connect(host='172.16.10.89',user='root',passwd='xxx123xxx',database='minor')
cursor = db.cursor()
query=("select *from student_data")
df = pd.read_sql_query(query,db)
print(df.head(10))
'''



def check_captcha():
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_TextBox1"]').clear()
    lst = list()
    lst.clear()
    images = driver.find_elements_by_tag_name('img')
    for image in images:
        a = image.get_attribute('src')
        lst.append(a)
    src = lst[1]
    response = requests.get(src)
    if response.status_code == 200:
        with open("sample.jpg", 'wb') as f:
            f.write(response.content)
    
    img = cv2.imread('sample.jpg', cv2.IMREAD_ANYCOLOR)
    
    text = pytesseract.image_to_string(img)
    text = text.replace(" ", "").upper()
    driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_TextBox1"]').send_keys(text)
    time.sleep(5)
    try:
        driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_btnviewresult"]').click()
        
    except:
        time.sleep(3)
    try:
        alert = driver.switch_to_alert
        driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_btnviewresult"]').click()
        check_captcha()
    except:
        get_result()
    
    
    
    
global sno
sno=1
global flag  
flag=1
def head():
    time.sleep(1)
    subject=[]
    subject.append("S.No.")
    subject.append("Roll No.")
    subject.append("Name")

    no_of_subjects=int(count_subjects.get())

    for i in range(no_of_subjects):
        sub=driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_pnlGrading"]/table/tbody/tr[3]/td/table['+str(i+2)+']/tbody/tr/td[1]')
        subject.append(sub.text)
        
    subject.append("SGPA")
    subject.append("CGPA")
    subject.append("RESULT")


    global flag
    flag=0
    with open('result.csv', 'w+', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(subject)
    
def get_result():
    
    
    det=[]
    try:
        alert = driver.switch_to_alert
        alert.accept()
        time.sleep(3)
        driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_btnReset"]').click()
        
        check_captcha()
    except:pass
    try:
        time.sleep(5)
        
        rno=driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_lblRollNoGrading"]')
        name=driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_lblNameGrading"]')    
        sgpa=driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_lblSGPA"]')    
        cgpa=driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_lblcgpa"]')    
        result=driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_lblResultNewGrading"]')
        global sno
        det.append(str(sno))
        
        sno+=1
        det.append(rno.text)
        det.append(name.text)
    
        no_of_subjects=int(count_subjects.get())
        if(flag==1):
            head()
        for i in range(no_of_subjects):
            grade=driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_pnlGrading"]/table/tbody/tr[3]/td/table['+str(i+2)+']/tbody/tr/td[4]')
            det.append(grade.text)
        det.append(sgpa.text)
        det.append(cgpa.text)
        det.append(result.text)
       
        with open('result.csv', 'a+', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(det)
       
        driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_btnReset"]').click()
        print(str(det[1])+"  "+str(det[2]))
    except:
        try:
            driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_btnReset"]').click()
        except:
        
            check_captcha()
        
def sel():
    
    if flag_csv==1: 

        global driver
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
        #chrome_options.add_argument("window-size=1024x768")
        chrome_options.add_argument("--no-sandbox")
        
        driver= webdriver.Chrome(executable_path='./chromedriver')#,options=chrome_options)
        driver.get('http://result.rgpv.ac.in/result/programselect.aspx?id=$%')
        driver.find_element_by_xpath('//*[@id="radlstProgram_'+str(var.get())+'"]').click()
        
        with open(str(tab1.filename), 'r') as f:
            data = csv.reader(f,delimiter=',')
            roll=list(data)
        
        for num in range(len(roll)):
            driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_txtrollno"]').send_keys(roll[num][0])
            driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_drpSemester"]/option['+str(num_sem.get())+']').click()
            check_captcha()
        driver.quit()
        global df
        df=pd.read_csv('result.csv')
        label6 = tk.LabelFrame(tab1, text="Generate Sheet:")
        label6.config(font=("Times New Roman", 26))
        label6.pack(fill="both")
        global str_cb
        str_cb = askstring("Sheet Details", "Course/Branch:")
        global str_ys
        str_ys = askstring("Sheet Details", "Year/Sem :")
        global str_b
        str_b = askstring("Sheet Details", "Batch :")
        global str_d
        str_d = askstring("Sheet Details", "Date :")
        button_sheet = tk.Button(label6,text ="Result Sheet", command = sheet1)
        button_sheet.config(font=("Times New Roman", 22))
        button_sheet.pack()
        button_main = tk.Button(label6,text ="Result Analysis", command = result_sheet)
        button_main.config(font=("Times New Roman", 22))
        button_main.pack()

    else:
        tk.messagebox.showinfo("Alert!", "Invalid Data Entry")

        
    
    
    
def sheet1():
    
    global wb
    wb=openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title="Sheet 1"
    sheet1['C8']=str_cb
    sheet1['C8'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['C8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1['G8']=str_ys
    sheet1['G8'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['G8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1['L8']=str_b
    sheet1['L8'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['L8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1['P8']=str_d
    sheet1['P8'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['P8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1.merge_cells('A1:P1')
    sheet1['A1'] = 'IPS Academy, Indore'
    sheet1['A1'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1.merge_cells('A2:P2')
    sheet1['A2'] = 'Institute of Engineering & Science'
    sheet1['A2'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)    
    sheet1.merge_cells('A3:P3')
    sheet1['A3'] = 'Department of Computer Science & Engineering'
    sheet1['A3'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['A3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1.merge_cells('A4:P4')
    sheet1['A4'] = 'Rajendra Nagar, A.B. Road, Indore 452012'
    sheet1['A4'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['A4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1.merge_cells('A5:P5')
    sheet1['A5'] = 'Tele fax0731-4014602 Phone: 4014645, 4014652'
    sheet1['A5'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['A5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1.merge_cells('A6:P6')
    sheet1['A6'] = 'E-mail: hod.compsc@ipsacademy.org,officecse@ipsacademy.org'
    sheet1['A6'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['A6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1.merge_cells('A7:P7')
    sheet1['A7'] = 'RESULT ANALYSIS'
    sheet1['A7'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['A7'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1.merge_cells('A8:B8')
    sheet1['A8']='Course/Branch:'
    sheet1['A8'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet1['A8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1['F8']='Year/Sem :'
    sheet1['F8'].font=Font(bold=True,name='Times New Roman',size=12)
    sheet1['F8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1['K8']='Batch :'
    sheet1['K8'].font=Font(bold=True,name='Times New Roman',size=12)
    sheet1['K8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet1['O8']='Date :'
    sheet1['O8'].font=Font(bold=True,name='Times New Roman',size=12)
    sheet1['O8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    img = openpyxl.drawing.image.Image('ips.jpg')
    img.width=100
    img.height=100
    sheet1.add_image(img,'D2')
    sheet1.column_dimensions['B'].width =15
    sheet1.column_dimensions['C'].width =30
    sheet1.column_dimensions['D'].width =13
    sheet1.column_dimensions['E'].width =13
    sheet1.column_dimensions['F'].width =13
    sheet1.column_dimensions['G'].width =13
    sheet1.column_dimensions['H'].width =13
    sheet1.column_dimensions['I'].width =13
    sheet1.column_dimensions['J'].width =13
    sheet1.column_dimensions['K'].width =13
    sheet1.column_dimensions['L'].width =13
    sheet1.column_dimensions['M'].width =13

    if count_subjects.get()=="11":
        sheet1.column_dimensions['N'].width =13
        sheet1.column_dimensions['O'].width =10
        sheet1.column_dimensions['P'].width =10
        sheet1.column_dimensions['Q'].width =25
    elif count_subjects.get()=="10":   
        sheet1.column_dimensions['N'].width =10
        sheet1.column_dimensions['O'].width =10
        sheet1.column_dimensions['P'].width =25
   
    rows = range(9, 100)
    columns = range(1, 20)
    for row in rows:
        for col in columns:
            sheet1.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet1.cell(row,col).font=Font(name='Times New Roman',size="12")
    
    r=range(10,100)
    c=range(3,4)
    for row in r:
        for col in c:
            sheet1.cell(row,col).alignment = Alignment(horizontal='left')
    
    dit()
    global csv_directory
    global str_filename
    wb.save(csv_directory +'/'+str_filename+'.xlsx')
    os.chmod(csv_directory +'/'+str_filename+'.xlsx', 0o777)
    

    wb.close()
    global df
    df1=df.copy()
    print(df1)
    global writer1
    writer1 = pd.ExcelWriter(csv_directory +'/'+str_filename+'.xlsx', engine='openpyxl')
    writer1.book = load_workbook(csv_directory +'/'+str_filename+'.xlsx')
    writer1.sheets = dict((ws1.title, ws1) for ws1 in writer1.book.worksheets)
    
    df1.to_excel(writer1, sheet_name='Sheet 1', index=False,startrow=8)
    
    writer1.save()
    
    writer1.close()

    
    tk.messagebox.showinfo("Alert!", "Sheet Generated")


def result_sheet():
    sort5_df=df.sort_values(by=['CGPA'], inplace=False, ascending=False).head(5)
    sort5_df=sort5_df[['Name','CGPA']] 
    count_row = df.shape[0]
    count_pass1=(df['RESULT']=='PASS').sum()
    count_pass2=(df['RESULT']=='PASS WITH GRACE').sum()
    count_pass=count_pass1+count_pass2
    pass_per=(count_pass/count_row)*100
    count_hons=(df['CGPA']>=7.5).sum()
    count_1div=((df['CGPA']>=6.5) & (df['CGPA']<7.5)).sum()
    count_2div=(df['CGPA']<6.5).sum()

    mapping = {'A+': 10, 'A': 9,'B+':8,'B':7,'C+':6,'C':5,'C##':5,'D##':4,'D':4,'F':3 ,'F (ABS)':3}
    sub_ls=list(df.columns)
    sub_t=[]
    for sub in sub_ls:
        if int(sub.find("[T]"))>0:
            sub_t.append(sub)
    try:
        for sub in sub_t:
            #print(sub)
            df.replace({sub: mapping},inplace=True)
    except:pass   
    count=[]
    sum_s=[]
    for sub in sub_t:
        count.append(int(df[(df[sub] > 3)].shape[0]))
        sum_s.append(df[sub].mean()*10)
    per_sub=[]
    for per in count:
        i=(per/count_row)*100
        per_sub.append(i)
    plt.close()
    plt.cla()
    plt.bar(sub_t, per_sub)
    plt.ylim(0,100)
    #plt.set_title("Pass percentage of students")
    plt.savefig('fig1.jpg')
    plt.close()
    plt.cla()
    plt.bar(sub_t,sum_s)
    plt.ylim(0,100)
    plt.savefig('fig2.jpg')
    
    global wb_main
    wb_main=openpyxl.Workbook()
    sheet2 = wb_main.active
    sheet2.title="Result Sheet"
    sheet2.merge_cells('A1:E1')
    sheet2['A1'] = 'Result Analysis'
    sheet2['A1'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet2['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet2.merge_cells('A2:E2')
    sheet2['A2'] = 'Department of Computer Science & Engineering'
    sheet2['A2'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet2['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet2.merge_cells('A5:C5')
    sheet2['A5'] = 'Top Five Students:'
    sheet2['A5'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet2['A5'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2.merge_cells('A3:E3')
    global str_cb
    global str_ys
    global str_b
    sheet2['A3']=str_cb+"  " + str_ys+"  " + str_b
    sheet2['A3'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet2['A3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet2['B6']="S.no."
    sheet2['B6'].font=Font(bold=True,name='Times New Roman',size="12")
    sheet2['B6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet2.column_dimensions['C'].width =30
    sheet2['B7']="1"
    sheet2['B7'].font=Font(name='Times New Roman',size="12")
    sheet2['B7'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet2['B8']="2"
    sheet2['B8'].font=Font(name='Times New Roman',size="12")
    sheet2['B8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet2['B9']="3"
    sheet2['B9'].font=Font(name='Times New Roman',size="12")
    sheet2['B9'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet2['B10']="4"
    sheet2['B10'].font=Font(name='Times New Roman',size="12")
    sheet2['B10'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet2['B11']="5"
    sheet2['B11'].font=Font(name='Times New Roman',size="12")
    sheet2['B11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet2.merge_cells('A13:C13')
    sheet2.merge_cells('A14:C14')
    sheet2.merge_cells('A15:C15')
    sheet2.merge_cells('A16:C16')
    sheet2.merge_cells('A17:C17')
    sheet2.merge_cells('A18:C18')
    sheet2.merge_cells('A21:C21')
    sheet2.merge_cells('A35:C35')
    sheet2['A13']='     Pass Percentage'
    sheet2['A13'].font=Font(name='Times New Roman',size="12")
    sheet2['A13'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['A14']='     Total Students Appearing'
    sheet2['A14'].font=Font(name='Times New Roman',size="12")
    sheet2['A14'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['A15']='     No. of students pass'
    sheet2['A15'].font=Font(name='Times New Roman',size="12")
    sheet2['A15'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['A16']='     No. of student passed with Hons.'
    sheet2['A16'].font=Font(name='Times New Roman',size="12")
    sheet2['A16'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['A17']='     No. of students passed in I Div.'
    sheet2['A17'].font=Font(name='Times New Roman',size="12")
    sheet2['A17'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['A18']='     No. of students passed in II Div.'
    sheet2['A18'].font=Font(name='Times New Roman',size="12")
    sheet2['A18'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['D13']=str(pass_per)+str("%") 
    sheet2['D13'].font=Font(name='Times New Roman',size="12")
    sheet2['D13'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['D14']=count_row
    sheet2['D14'].font=Font(name='Times New Roman',size="12")
    sheet2['D14'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['D15']=count_pass
    sheet2['D15'].font=Font(name='Times New Roman',size="12")
    sheet2['D15'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['D16']=count_hons
    sheet2['D16'].font=Font(name='Times New Roman',size="12")
    sheet2['D16'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['D17']=count_1div
    sheet2['D17'].font=Font(name='Times New Roman',size="12")
    sheet2['D17'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['D18']=count_2div
    sheet2['D18'].font=Font(name='Times New Roman',size="12")
    sheet2['D18'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    sheet2['A21']="Pass percentage of students:-"
    sheet2['A21'].font=Font(name='Times New Roman',size="12")
    sheet2['A21'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet2['A35']="Average percentage of marks:-"
    sheet2['A35'].font=Font(name='Times New Roman',size="12")
    sheet2['A35'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    img = openpyxl.drawing.image.Image('fig1.jpg')
    img.width=350
    img.height=250
    sheet2.add_image(img,'A22')
    img = openpyxl.drawing.image.Image('fig2.jpg')
    img.width=350
    img.height=250
    sheet2.add_image(img,'A36')
    
    r=range(7,12)
    c=range(2,5)
    for row in r:
        for col in c:
            sheet2.cell(row,col).font=Font(name='Times New Roman',size="12")
            if col==3:
                sheet2.cell(row, col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                sheet2.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

     
            
    
    dit()
    global csv_directory
    global str_filename
    wb_main.save(csv_directory +'/'+str_filename+'.xlsx')
    os.chmod(csv_directory +'/'+str_filename+'.xlsx', 0o777)
    
    writer2 = pd.ExcelWriter(csv_directory +'/'+str_filename+'.xlsx', engine='openpyxl')
    writer2.book = load_workbook(csv_directory +'/'+str_filename+'.xlsx')
    writer2.sheets = dict((ws2.title, ws2) for ws2 in writer2.book.worksheets)   
    
    sort5_df.to_excel(writer2, sheet_name='Result Sheet', index=False,startrow=5,startcol=2)
        
    writer2.save()
    
    writer2.close()
    
    tk.messagebox.showinfo("Alert!", "Result Sheet Generated")

    

def choose():
    tab1.filename =  filedialog.askopenfilename(title = "Select file",filetypes = (("csv files","*.csv"),("all files","*.*")))
    print("File path:"+str(tab1.filename))
    global flag_csv
    flag_csv=1

def dit():
    global csv_directory
    tab1.directory =  filedialog.askdirectory()
    csv_directory=str(tab1.directory)
    print("Save directory:"+csv_directory)
    global str_filename
    str_filename = askstring("File Details", "Enter File Name:")
    print('File Name:'+str_filename)  


if __name__=='__main__':
    global flag_csv
    flag_csv=0
    
    global df_pre
    global regressor_2,X_test_2,y_test_2
    global regressor_3,X_test_3,y_test_3
    global regressor_4,X_test_4,y_test_4
    global regressor_5,X_test_5,y_test_5
    global regressor_6,X_test_6,y_test_6
    global regressor_7,X_test_7,y_test_7
    global regressor_8,X_test_8,y_test_8
    df_pre=pd.read_csv("export_df.csv")
    
    #sem2
    df2=df_pre.copy()
    df2.dropna(subset = ['1','2'], inplace=True)
    #print(df2)
    X_2=df2[['1']]
    y_2=df2['2']
    
    X_train_2, X_test_2, y_train_2, y_test_2 = train_test_split(X_2, y_2, test_size=0.33, random_state=100) 
    regressor_2 = RandomForestRegressor(n_estimators = 100, random_state = 0) 
    regressor_2.fit(X_train_2, y_train_2)
    #sem3
    df3=df_pre.copy()
    df3.dropna(subset = ['1','2','3'], inplace=True)
    #print(df3)
    X_3=df3[['1','2']]
    y_3=df3['3']
    X_train_3, X_test_3, y_train_3, y_test_3 = train_test_split(X_3, y_3, test_size=0.33, random_state=100)
    regressor_3 = RandomForestRegressor(n_estimators = 100, random_state = 0) 
    regressor_3.fit(X_train_3, y_train_3)
    #sem4
    df4=df_pre.copy()
    df4.dropna(subset = ['1','2','3','4'], inplace=True)
    #print(df4)
    X_4=df4[['1','2','3']]
    y_4=df4['4']
    X_train_4, X_test_4, y_train_4, y_test_4 = train_test_split(X_4, y_4, test_size=0.33, random_state=100)
    regressor_4 = RandomForestRegressor(n_estimators = 100, random_state = 0) 
    regressor_4.fit(X_train_4, y_train_4)
    #sem5
    df5=df_pre.copy()
    df5.dropna(subset = ['1','2','3','4','5'], inplace=True)
    #print(df5)
    X_5=df5[['1','2','3','4']]
    y_5=df5['5']
    X_train_5, X_test_5, y_train_5, y_test_5 = train_test_split(X_5, y_5, test_size=0.33, random_state=100)
    regressor_5 = RandomForestRegressor(n_estimators = 100, random_state = 0) 
    regressor_5.fit(X_train_5, y_train_5)
    #sem6
    df6=df_pre.copy()
    df6.dropna(subset = ['1','2','3','4','5','6'], inplace=True)
    #print(df6)
    X_6=df6[['1','2','3','4','5']]
    y_6=df6['6']
    X_train_6, X_test_6, y_train_6, y_test_6 = train_test_split(X_6, y_6, test_size=0.33, random_state=100)
    regressor_6 = RandomForestRegressor(n_estimators = 100, random_state = 0) 
    regressor_6.fit(X_train_6, y_train_6)
    #sem7
    df7=df_pre.copy()
    df7.dropna(subset = ['1','2','3','4','5','6','7'], inplace=True)
    #print(df7)
    X_7=df7[['1','2','3','4','5','6']]
    y_7=df7['7']
    X_train_7, X_test_7, y_train_7, y_test_7 = train_test_split(X_7, y_7, test_size=0.33, random_state=100)
    regressor_7 = RandomForestRegressor(n_estimators = 100, random_state = 0) 
    regressor_7.fit(X_train_7, y_train_7)
    #sem8
    df8=df_pre.copy()
    df8.dropna(subset = ['1','2','3','4','5','6','7','8'], inplace=True)
    #print(df8)
    X_8=df8[['1','2','3','4','5','6','7']]
    y_8=df8['8']
    X_train_8, X_test_8, y_train_8, y_test_8 = train_test_split(X_8, y_8, test_size=0.33, random_state=100)
    regressor_8 = RandomForestRegressor(n_estimators = 100, random_state = 0) 
    regressor_8.fit(X_train_8, y_train_8)
    
def predict():
    global prediction_fig
    try:
        prediction_fig.destroy()
    except:pass
    global ls_sgpa , sgpa_curr ,strrno,flag_rno
    ls_sgpa=[]
    sgpa_curr=[]
    if (len(rno.get().strip())!=0):
        strrno=rno.get().upper()
        flag_rno=0
        print("Enrollment No.: "+strrno)
        dfrno=df_pre[df_pre['rno']==strrno]
        #print(dfrno['rno'].values)
        if (dfrno['rno'].values==strrno):
            flag_rno=1
            ls=[]
            #print(dfrno.values)
            for i in range(8):
                if (dfrno[str(i+1)].values>0):
                    ls.append(dfrno[str(i+1)].values)
                    ls_sgpa.append(ls[i][0])    
            sgpa_curr=ls_sgpa.copy()
        else:
             tk.messagebox.showinfo("ERROR!", "Data not found for "+strrno)
    else:
        flag_rno=0
        s1=texts1.get()
        s2=texts2.get()
        s3=texts3.get()
        s4=texts4.get()
        s5=texts5.get()
        s6=texts6.get()
        s7=texts7.get()
        s8=texts8.get()
        
        ls_sgpa=[]
        sgpa_curr=[]
        ls_s=[]
        ls_s.append(s1)
        ls_s.append(s2)
        ls_s.append(s3)
        ls_s.append(s4)
        ls_s.append(s5)
        ls_s.append(s6)
        ls_s.append(s7)
        ls_s.append(s8)
        
        for i in ls_s:
            if len(i.strip())>0:
                try:
                    if 0<float(i)<=10:
                        ls_sgpa.append(float(i))
                    else:
                        tk.messagebox.showinfo("ERROR!", "Enter Valid Data")
                except:
                    tk.messagebox.showinfo("ERROR!", "Enter Valid Data")
        sgpa_curr=ls_sgpa.copy()        
        
    if (len(ls_sgpa)==8):
        sem()
    elif(len(ls_sgpa)==7):
        sem8()
    elif(len(ls_sgpa)==6):
        sem7()
    elif(len(ls_sgpa)==5):
        sem6()
    elif(len(ls_sgpa)==4):
        sem5()
    elif(len(ls_sgpa)==3):
        sem4()
    elif(len(ls_sgpa)==2):
        sem3()
    elif(len(ls_sgpa)==1):
        sem2()
def sem():
    #print("sem")
    plt.close()
    plt.cla()
    plt.plot(['Sem 1','Sem 2','Sem 3','Sem 4','Sem 5','Sem 6','Sem 7','Sem 8'],ls_sgpa ,c='orange',marker='o',lw=2.5 ,label="Current")
    
    plt.legend(framealpha=1,frameon=True)
    plt.ylim(0,10)
    plt.ylabel('SGPA')
    plt.tight_layout()
    plt.savefig('prediction1.jpg')
    os.chmod('prediction1.jpg', 0o777)
    #plt.show()
    

    
    pre_img=Image.open('prediction1.jpg')
    pre_img=pre_img.resize((550,350),Image.ANTIALIAS)
    pre_image=ImageTk.PhotoImage(pre_img)
    

    prediction_fig=tk.Label(tab2,image=pre_image)
    prediction_fig.image=pre_image
    #prediction_fig.place(x=800,y=100)
    prediction_fig.grid(row=1,rowspan=25,column=4,padx=150)
            
def sem2():
    predict_2 =regressor_2.predict([ls_sgpa,])
    predict_2=round(predict_2[0],2)
    #print(predict_6)
    str_accuracy=str(regressor_2.score(X_test_2,y_test_2)*100)
    print("Sem 2")
    print("Prediction: "+ str(predict_2)+" SGPA")
    print("Prediction Accuracy Rate: "+str_accuracy[:5]+"%")
    ls_sgpa.append(predict_2)
   
    sem3()
    
def sem3():
    predict_3 =regressor_3.predict([ls_sgpa,])
    predict_3=round(predict_3[0],2)
    #print(predict_6)
    str_accuracy=str(regressor_3.score(X_test_3,y_test_3)*100)
    print("Sem 3")
    print("Prediction: "+ str(predict_3)+" SGPA")
    print("Prediction Accuracy Rate: "+str_accuracy[:5]+"%")
    ls_sgpa.append(predict_3)

    sem4()
def sem4():
    predict_4 =regressor_4.predict([ls_sgpa,])
    predict_4=round(predict_4[0],2)
    #print(predict_6)
    str_accuracy=str(regressor_4.score(X_test_4,y_test_4)*100)
    print("Sem 4")
    print("Prediction: "+ str(predict_4)+" SGPA")
    print("Prediction Accuracy Rate: "+str_accuracy[:5]+"%")
    ls_sgpa.append(predict_4)

    sem5()
def sem5():
    predict_5 =regressor_5.predict([ls_sgpa,])
    predict_5=round(predict_5[0],2)
    #print(predict_6)
    str_accuracy=str(regressor_5.score(X_test_5,y_test_5)*100)
    print("Sem 5")
    print("Prediction: "+ str(predict_5)+" SGPA")
    print("Prediction Accuracy Rate: "+str_accuracy[:5]+"%")
    ls_sgpa.append(predict_5)


    sem6()
def sem6():
    predict_6 =regressor_6.predict([ls_sgpa,])
    predict_6=round(predict_6[0],2)
    #print(predict_6)
    str_accuracy=str(regressor_6.score(X_test_6,y_test_6)*100)
    print("Sem 6")
    print("Prediction: "+ str(predict_6)+" SGPA")
    print("Prediction Accuracy Rate: "+str_accuracy[:5]+"%")
    ls_sgpa.append(predict_6)
    
    sem7()
def sem7():
    predict_7 =regressor_7.predict([ls_sgpa,])
    predict_7=round(predict_7[0],2)
    #print(predict_7)
    str_accuracy=str(regressor_7.score(X_test_7,y_test_7)*100)
    print('Sem 7')
    print("Prediction: "+ str(predict_7)+" SGPA")
    print("Prediction Accuracy Rate: "+str_accuracy[:5]+"%")
    ls_sgpa.append(predict_7)
    
    sem8()
    
def sem8():
    sem=[]
    for i in range(len(sgpa_curr)):
        sem.append("Sem "+str(i+1))
    predict_8 =regressor_8.predict([ls_sgpa,])
    predict_8=round(predict_8[0],2)
    #print(predict_6)
    str_accuracy=str(regressor_8.score(X_test_8,y_test_8)*100)
    print('Sem 8')
    print("Prediction: "+ str(predict_8)+" SGPA")
    print("Prediction Accuracy Rate: "+str_accuracy[:5]+"%")
    ls_sgpa.append(predict_8)
    #print(sgpa_curr)
    if flag_rno==1:
        #print(strrno[6:8])
        rno=strrno[6:8]
        df=df_pre.copy()
        df_rno=df[df['rno'].str.contains('0808CS'+rno)|df['rno'].str.contains('0808CO'+rno)|df['rno'].str.contains('0808CI'+rno)|df['rno'].str.contains('0808CT'+rno)]
        lsavg=df_rno[['1','2','3','4','5','6','7','8']].mean(axis=0)
        plt.close()
        plt.cla()
        plt.plot(sem,lsavg[:len(sgpa_curr)] ,marker='.',c='green',lw=2.5,label='Batch Avg')
    else:
        plt.close()
        plt.cla()
    plt.plot(['Sem 1','Sem 2','Sem 3','Sem 4','Sem 5','Sem 6','Sem 7','Sem 8'],ls_sgpa ,marker='*',lw=2.0 ,label="Prediction")
    plt.plot(sem,sgpa_curr ,marker='o',lw=2.5,label='Current')
    plt.legend(framealpha=1,frameon=True)
    plt.ylim(0,10)
    plt.ylabel('SGPA')
    plt.tight_layout()
    plt.savefig('prediction1.jpg')
    os.chmod('prediction1.jpg', 0o777)
    #plt.show()
    

    
    pre_img=Image.open('prediction1.jpg')
    pre_img=pre_img.resize((550,350),Image.ANTIALIAS)
    pre_image=ImageTk.PhotoImage(pre_img)
    
    global prediction_fig
    prediction_fig=tk.Label(tab2,image=pre_image)
    prediction_fig.image=pre_image
    prediction_fig.grid(row=1,rowspan=25,column=4,padx=150)
       

   
root=tk.Tk()
frame = tk.Frame(root)
frame.pack()
root.title("RGPV EXAM RESULT ANALYSIS")     

tab_parent = ttk.Notebook(root)
tab1 = ttk.Frame(tab_parent)
tab2 = ttk.Frame(tab_parent)
tab_parent.add(tab1, text="Result Analysis")
tab_parent.add(tab2, text="Result Prediction")


var = tk.IntVar()
label = tk.LabelFrame(tab1, text="Select Course:")
label.config(font=("Times New Roman", 26))
label.pack(fill="both")
R1 = tk.Radiobutton(label, text="B.E.     ", variable=var, value=0)
R1.config(font=("Times New Roman", 18))
R1.pack( anchor = CENTER)
R2 = tk.Radiobutton(label, text="B.Tech", variable=var, value=1)
R2.config(font=("Times New Roman", 18))
R2.pack( anchor = CENTER)

label1 = tk.LabelFrame(tab1, text="Select Sem:")
label1.config(font=("Times New Roman", 26))
label1.pack(fill="both")
num_sem = tk.Spinbox(label1, from_=1, to=8)
num_sem.config(font=("Times New Roman", 18))
num_sem.pack(anchor = CENTER )

label4=tk.LabelFrame(tab1,text="Choose File (for Enrollment No.)")
label4.config(font=("Times New Roman", 26))
label4.pack(fill="both")
button = tk.Button(label4,text ="Open", command = choose)
button.config(font=("Times New Roman", 18))
button.pack(anchor = CENTER )


label_sub = tk.LabelFrame(tab1, text="Enter count of Subjects:")
label_sub.pack(fill="both")
label_sub.config(font=("Times New Roman", 26))
count_subjects = tk.Spinbox(label_sub, from_=10, to=11)
count_subjects.config(font=("Times New Roman", 18))
count_subjects.pack(anchor = CENTER)

button = tk.Button(tab1,text ="Submit", command = sel)
button.config(font=("Times New Roman", 26))
button.pack(anchor = CENTER)

label_rno = tk.Label(tab2, text="Enrollment No.:")
label_rno.config(font=("Times New Roman", 26))
rno=tk.Entry(tab2)
rno.config(font=("Times New Roman", 20))

labelor = tk.Label(tab2, text="OR")
labelor.config(font=("Times New Roman", 26))

label_sem = tk.Label(tab2, text="Individual SGPA's:-")
label_sem.config(font=("Times New Roman", 24))


labels1 = tk.Label(tab2, text="Sem 1:")
texts1=tk.Entry(tab2)
labels2 = tk.Label(tab2, text="Sem 2:")
texts2=tk.Entry(tab2)
labels3 = tk.Label(tab2, text="Sem 3:")
texts3=tk.Entry(tab2)
labels4 = tk.Label(tab2, text="Sem 4:")
texts4=tk.Entry(tab2)
labels5 = tk.Label(tab2, text="Sem 5:")
texts5=tk.Entry(tab2)
labels6 = tk.Label(tab2, text="Sem 6:")
texts6=tk.Entry(tab2)
labels7 = tk.Label(tab2, text="Sem 7:")
texts7=tk.Entry(tab2)
labels8 = tk.Label(tab2, text="Sem 8:")
texts8=tk.Entry(tab2)
button_generate = tk.Button(tab2,text ="Generate", command = predict)
button_generate.config(font=("Times New Roman", 26))
def clear_text():
    rno.delete(0,'end')
    texts1.delete(0, 'end')
    texts2.delete(0, 'end')
    texts3.delete(0, 'end')
    texts4.delete(0, 'end')
    texts5.delete(0, 'end')
    texts6.delete(0, 'end')
    texts7.delete(0, 'end')
    texts8.delete(0, 'end')
    prediction_fig.destroy()
    rno.focus()
clear_button = tk.Button(tab2,text="Clear", command=clear_text)
clear_button.config(font=("Times New Roman", 26))



labels1.config(font=("Times New Roman", 20))
texts1.config(font=("Times New Roman", 20))
labels2.config(font=("Times New Roman", 20))
texts2.config(font=("Times New Roman", 20))
labels3.config(font=("Times New Roman", 20))
texts3.config(font=("Times New Roman", 20))
labels4.config(font=("Times New Roman", 20))
texts4.config(font=("Times New Roman", 20))
labels5.config(font=("Times New Roman", 20))
texts5.config(font=("Times New Roman", 20))
labels6.config(font=("Times New Roman", 20))
texts6.config(font=("Times New Roman", 20))
labels7.config(font=("Times New Roman", 20))
texts7.config(font=("Times New Roman", 20))
labels8.config(font=("Times New Roman", 20))
texts8.config(font=("Times New Roman", 20))



label_rno.grid(row=0,column=0,pady=10,padx=30)
rno.grid(row=0,column=1)
labelor.grid(row=1,column=0,pady=5,columnspan=2)
label_sem.grid(row=2,column=0,padx=25)
labels1.grid(row=3, column=0)
labels2.grid(row=4, column=0)
labels3.grid(row=5, column=0)
labels4.grid(row=6, column=0)
labels5.grid(row=7, column=0)
labels6.grid(row=8, column=0)
labels7.grid(row=9, column=0)
labels8.grid(row=10, column=0)
texts1.grid(row=3,column=1)
texts2.grid(row=4,column=1)
texts3.grid(row=5,column=1)
texts4.grid(row=6,column=1)
texts5.grid(row=7,column=1)
texts6.grid(row=8,column=1)
texts7.grid(row=9,column=1)
texts8.grid(row=10,column=1)
clear_button.grid(row=12,column=0,pady=15)
button_generate.grid(row=12,column=1,columnspan=2,pady=15)


tab_parent.pack(expand=1,fill="both")
root.mainloop()
